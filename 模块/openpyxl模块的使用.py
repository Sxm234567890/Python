#创建工作簿
'''
import weather
import openpyxl
html=weather.get_html()
lst=weather.parse_html(html)
#创建一个新的excel工作簿
workbook=openpyxl.Workbook()  #创建对象
sheet=workbook.create_sheet('景区天气')
for item in lst:
    sheet.append(item)

workbook.save('景区天气.xlsx')
'''

#查看工作簿
import openpyxl
#打开工作簿
workbook=openpyxl.load_workbook('景区天气.xlsx')
#选择要操作得工作表
sheet=workbook['景区天气']
#Biaoshi二维列表
lst=[]  #存储行得数据
for row in sheet.rows:
    sublst=[]#存储单元格数据
    for cell in row:
        sublst.append(cell.value) #存储列的数据
    lst.append(sublst)

for item in lst:
    print(item)







